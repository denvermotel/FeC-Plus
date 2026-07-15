#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# FeC-Plus - v0.03 alpha
"""
fec_utility.py - Utility di riepilogo: export Excel «Elenco fatture».

Libreria pura, separata da `fec_download.py` (che resta il motore di download):
riceve un `AuthResult` già autenticato (vedi `ade_auth.autentica`) e genera un
file .xlsx di riepilogo con una riga per fattura e colonne pivot per aliquota
IVA / codice natura (formato ricalcato dallo userscript FE-Utility,
https://github.com/denvermotel/fe-utility, che però scrapa il DOM del portale:
qui si usano le API JSON di consultazione già note a `fec_download.py`).

Funzioni pubbliche:
    elenco_fatture_excel(auth, cf_cliente, piva, tipo, dal, al,
                         dest_dir=None, sottocartella=True,
                         control=None, log=print) -> str   # percorso xlsx
    elenco_corrispettivi_excel(auth, cf_cliente, piva, dal, al,
                               ...come sopra) -> list[str] # un xlsx per matricola

`tipo` (vedi TIPI_ELENCO): "emesse" | "ricevute_ricezione" | "ricevute_emissione"
| "trans_emesse" | "trans_ricevute". Date `dal`/`al` in formato GGMMAAAA; i
periodi > 3 mesi vengono spezzati internamente (fec_queue.spezza_periodo, con la
guardia dei 12 mesi) e le liste concatenate: l'output è comunque UN solo file.

Dipendenza opzionale `openpyxl` (ruolo "excel" in fec_deps), import lazy solo
alla scrittura del file. Per i corrispettivi la P.IVA è obbligatoria (fa parte
degli URL AdE); vedi la sezione Corrispettivi più sotto per gli endpoint.

Per i nomi campo JSON delle fatture (fatture emesse, ma il concetto vale anche
per gli altri tipi): la LISTA (`denominazione/piva/cf<Cliente|Fornitore>`) è la
fonte primaria per l'anagrafica della controparte, sempre disponibile anche
quando il dettaglio fallisce; il dettaglio (endpoint `rs/fatture/dettaglio/...`)
può tornare HTTP 200 con un corpo vuoto/placeholder e un avviso `messages[]`
(«Dati momentaneamente non disponibili.»): errore temporaneo lato AdE, non di
rete, per cui si ritenta una volta prima di arrendersi (vedi
`_dettaglio_fattura`/`_avviso_dettaglio`).
"""

from __future__ import annotations

__version__ = "0.03 alpha"

import json
import os
import time

import requests

from ade_auth import AuthResult, IVASERVIZI, unix_time
from fec_download import (_cartella, Controllo, DownloadError,  # noqa: F401
                          HTTP_TIMEOUT)
from fec_queue import spezza_periodo


class NessunDato(DownloadError):
    """Nessuna fattura/corrispettivo nel periodo: nessun file generato. Distinta da
    DownloadError così i chiamanti che ciclano su più sotto-periodi (Excel per
    trimestre/mese) possono saltare i periodi vuoti senza abortire il lavoro."""


def spezza_granularita(dal: str, al: str, granularita: str,
                       fmt: str = "%d%m%Y") -> list[tuple[str, str]]:
    """
    Spezza [dal, al] in sotto-periodi per generare PIÙ file Excel (scelta utente
    per l'anno intero): `granularita` = "unico" (default, nessuna divisione),
    "trimestre" (blocchi allineati ai trimestri solari) o "mese" (mesi solari).
    Primo/ultimo blocco tagliati su dal/al. Diversa da `fec_queue.spezza_periodo`
    (finestre mobili da N mesi per il limite AdE): qui i confini sono solari,
    come si aspetta chi archivia un file per trimestre/mese.
    """
    if granularita in (None, "", "unico"):
        return [(dal, al)]
    if granularita not in ("trimestre", "mese"):
        raise DownloadError(f"Granularità sconosciuta: {granularita!r} "
                            "(valide: unico, trimestre, mese).")
    from datetime import date, datetime, timedelta
    from dateutil.relativedelta import relativedelta
    try:
        d0 = datetime.strptime(dal, fmt).date()
        a0 = datetime.strptime(al, fmt).date()
    except (ValueError, TypeError):
        raise DownloadError("Date non valide per la divisione in più file.")
    if d0 > a0:
        raise DownloadError("La data inizio è successiva alla data fine.")

    blocchi: list[tuple[str, str]] = []
    start = d0
    while start <= a0:
        if granularita == "trimestre":
            primo_mese = ((start.month - 1) // 3) * 3 + 1
            fine = (date(start.year, primo_mese, 1)
                    + relativedelta(months=3) - timedelta(days=1))
        else:
            fine = (date(start.year, start.month, 1)
                    + relativedelta(months=1) - timedelta(days=1))
        fine = min(fine, a0)
        blocchi.append((start.strftime(fmt), fine.strftime(fmt)))
        start = fine + timedelta(days=1)
    return blocchi


# ─────────────────────────────────────────────────────────────────────────────
# Endpoint di consultazione (gli stessi di fec_download, sola lettura elenco)
# ─────────────────────────────────────────────────────────────────────────────

# tipo → (frammento URL lista, etichetta per log/nome file)
TIPI_ELENCO: dict[str, tuple[str, str]] = {
    "emesse":             ("fe/emesse/dal/{dal}/al/{al}",                     "emesse"),
    "ricevute_ricezione": ("fe/ricevute/dal/{dal}/al/{al}/ricerca/ricezione", "ricevute"),
    "ricevute_emissione": ("fe/ricevute/dal/{dal}/al/{al}/ricerca/emissione", "ricevute"),
    "trans_emesse":       ("ft/emesse/dal/{dal}/al/{al}",                     "trans_emesse"),
    "trans_ricevute":     ("ft/ricevute/dal/{dal}/al/{al}",                   "trans_ricevute"),
}


def _lista_fatture(auth: AuthResult, tipo: str, dal: str, al: str) -> list[dict]:
    """
    Elenco fatture (voci JSON grezze) per il blocco [dal, al] (GGMMAAAA).
    Stesse liste usate da `fec_download._scarica_da_lista`, qui senza scaricare i file.
    """
    try:
        frammento, _ = TIPI_ELENCO[tipo]
    except KeyError:
        raise DownloadError(f"Tipo elenco sconosciuto: {tipo!r} "
                            f"(validi: {', '.join(TIPI_ELENCO)}).")
    url = (f"{IVASERVIZI}/cons/cons-services/rs/"
           f"{frammento.format(dal=dal, al=al)}?v={unix_time()}")
    try:
        r = auth.session.get(url, headers=auth.headers, verify=False, timeout=HTTP_TIMEOUT)
    except requests.exceptions.RequestException as exc:
        raise DownloadError(
            f"Nessuna risposta dall'AdE per l'elenco fatture entro il tempo limite "
            f"({HTTP_TIMEOUT[1]}s): {exc}. Riprova; se persiste, riduci il periodo.")
    if r.status_code != 200:
        raise DownloadError(
            f"La richiesta dell'elenco fatture è fallita (HTTP {r.status_code}). "
            "La sessione potrebbe essere scaduta o l'utenza di lavoro non attiva.")
    try:
        data = r.json()
    except ValueError:
        raise DownloadError("Risposta non in formato JSON: probabile sessione "
                            "scaduta o redirect alla pagina di login.")
    return data.get("fatture") or []


def _avviso_dettaglio(dati: dict) -> str | None:
    """
    Testo del primo avviso AdE (`messages[]`, severity WARNING/ERROR) se il
    dettaglio è tornato HTTP 200 ma con corpo vuoto/placeholder, ad esempio
    `{"messages": [{"message": "Dati momentaneamente non disponibili.",
    "severity": "WARNING", ...}], "idFattura": "", ...}`: un errore temporaneo
    lato AdE distinto da un guasto HTTP/di rete. None se il dettaglio è valido.
    """
    for msg in (dati.get("messages") or []):
        if str(msg.get("severity", "")).upper() in ("WARNING", "ERROR"):
            return str(msg.get("message") or "Dati non disponibili")
    return None


def _dettaglio_fattura(auth: AuthResult, fattura_file: str, log=print,
                       tentativi: int = 2) -> dict | None:
    """
    Dettaglio JSON di una fattura (`fattura_file` = tipoInvio + idFattura), lo stesso
    endpoint usato da `fec_download._e_scartata_pa`. Contiene le righe di riepilogo
    IVA (imponibile/aliquota/imposta/natura). Ritenta se l'AdE risponde con un
    avviso «dati non disponibili» (vedi `_avviso_dettaglio`, transitorio: nella
    cattura di scoperta è bastato riaprire il dettaglio). Su errore persistente
    ritorna None (il chiamante logga e produce una riga con i soli dati di lista).
    """
    url = (f"{IVASERVIZI}/cons/cons-services/rs/fatture/dettaglio/{fattura_file}"
           f"?v={unix_time()}")
    for tentativo in range(1, tentativi + 1):
        try:
            r = auth.session.get(url, headers=auth.headers, verify=False,
                                 timeout=HTTP_TIMEOUT)
        except requests.exceptions.RequestException as exc:
            log(f"   ⚠️  Dettaglio {fattura_file} non ottenuto ({exc}); "
                "riga con i soli dati di elenco.")
            return None
        if r.status_code != 200:
            log(f"   ⚠️  Dettaglio {fattura_file}: HTTP {r.status_code}; "
                "riga con i soli dati di elenco.")
            return None
        try:
            dati = r.json()
        except ValueError:
            log(f"   ⚠️  Dettaglio {fattura_file}: risposta non JSON; "
                "riga con i soli dati di elenco.")
            return None
        avviso = _avviso_dettaglio(dati)
        if avviso is None:
            return dati
        if tentativo < tentativi:
            log(f"   ⚠️  Dettaglio {fattura_file}: {avviso} (AdE); ritento...")
            time.sleep(1.0)
    log(f"   ⚠️  Dettaglio {fattura_file} ancora non disponibile dopo {tentativi} "
        "tentativi; riga con i soli dati di elenco.")
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Parser difensivo (i nomi campo JSON dell'AdE vanno confermati col dump di
# discovery: qui si prova una rosa di alias e si degrada a vuoto se assenti)
# ─────────────────────────────────────────────────────────────────────────────

# Colonne fisse dell'Excel fatture (prima delle coppie pivot per aliquota).
# "ID Invio" = idInvio SDI (idFattura come fallback: identificativo interno AdE,
# non l'idInvio, usato solo se il primo non è disponibile); "Codice Fiscale"
# accanto a "Partita IVA" perché la controparte può avere solo uno dei due.
COLONNE_FISSE = ["Data", "N. Fattura", "ID Invio", "Tipo Documento",
                 "Cliente/Fornitore", "Partita IVA", "Codice Fiscale"]

# Tipi documento che rappresentano note di credito (importi da negare).
_TIPI_NOTA_CREDITO = ("TD04", "TD08")


def _campo(d: dict | None, *nomi: str, default=None):
    """Primo valore non-None tra gli alias `nomi` (match case-insensitive sulle chiavi)."""
    if not isinstance(d, dict):
        return default
    minuscole = {k.lower(): v for k, v in d.items()}
    for nome in nomi:
        val = minuscole.get(nome.lower())
        if val is not None:
            return val
    return default


def _importo(val) -> float:
    """
    Converte un importo AdE in float: gestisce numeri nativi, il formato grezzo
    `+000000000012,00`, separatori italiani `1.234,56` e stringhe vuote → 0.0.
    """
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace("€", "").strip()
    if not s:
        return 0.0
    segno = -1.0 if s.startswith("-") else 1.0
    s = s.lstrip("+-")
    if "," in s:                       # formato italiano: la virgola è il decimale
        s = s.replace(".", "").replace(",", ".")
    try:
        return segno * float(s)
    except ValueError:
        return 0.0


def _chiave_aliquota(riga_iva: dict) -> str | None:
    """
    Chiave pivot di una riga di riepilogo IVA: `"22%"` per le aliquote numeriche,
    il codice natura (`"N2.1"`) per le operazioni senza imposta, None se indeterminabile.
    """
    aliquota = _campo(riga_iva, "aliquotaIVA", "aliquota", "aliquotaIva")
    natura = _campo(riga_iva, "natura", "codiceNatura", "naturaOperazione")
    if aliquota is not None and str(aliquota).strip():
        n = _importo(aliquota)
        if n > 0 or not natura:
            return f"{n:g}%".replace(".", ",")     # 22.0 → "22%", 10.5 → "10,5%"
    if natura and str(natura).strip():
        return str(natura).strip().upper()
    return None


def _righe_riepilogo_iva(dettaglio: dict | None) -> list[dict]:
    """
    Trova nel JSON di dettaglio la lista delle righe di riepilogo IVA, senza conoscere
    la chiave contenitore: cerca ricorsivamente la prima lista di dict che abbia sia
    un campo imponibile sia un campo imposta/aliquota (struttura della tabella
    «Riepiloghi» della pagina Dettaglio fattura del portale).
    """
    def _sembra_riepilogo(lst) -> bool:
        return (isinstance(lst, list) and lst and isinstance(lst[0], dict)
                and _campo(lst[0], "imponibile", "imponibileImporto") is not None
                and (_campo(lst[0], "imposta", "aliquotaIVA", "aliquota",
                            "aliquotaIva") is not None))

    def _cerca(nodo):
        if _sembra_riepilogo(nodo):
            return nodo
        if isinstance(nodo, dict):
            for v in nodo.values():
                trovato = _cerca(v)
                if trovato is not None:
                    return trovato
        elif isinstance(nodo, list):
            for v in nodo:
                trovato = _cerca(v)
                if trovato is not None:
                    return trovato
        return None

    return _cerca(dettaglio) or []


def _e_nota_credito(tipo_doc: str) -> bool:
    t = (tipo_doc or "").upper()
    return any(td in t for td in _TIPI_NOTA_CREDITO) or "NOTA DI CREDITO" in t


def _data_it(val) -> str:
    """
    Normalizza una data eterogenea AdE in formato it `GG/MM/AAAA`: converte l'ISO
    `YYYY-MM-DD[THH:MM:SS]` (lista fatture, `timeRilevazione` corrispettivi),
    stacca l'eventuale orario da `DD/MM/YYYY HH:MM` (`dataAccoglienzaFile`
    corrispettivi) e lascia invariato un `DD/MM/YYYY` già pulito (dettaglio fatture).
    """
    s = str(val or "").strip()
    if "T" in s:
        s = s.split("T")[0]
    if len(s) >= 10 and s[4:5] == "-" and s[7:8] == "-":
        anno, mese, giorno = s[:10].split("-")
        return f"{giorno}/{mese}/{anno}"
    return s.split(" ")[0]


def _nome_persona(d: dict | None, suffisso: str) -> str:
    """Combina nome+cognome (persona fisica) per un ruolo ("Cliente"/"Fornitore"):
    il dettaglio li tiene separati (spesso con `denominazione<ruolo>` vuoto)."""
    nome = str(_campo(d, f"nome{suffisso}", default="") or "").strip()
    cognome = str(_campo(d, f"cognome{suffisso}", default="") or "").strip()
    return " ".join(p for p in (nome, cognome) if p)


def _dati_controparte(voce: dict, dettaglio: dict | None) -> tuple[str, str, str]:
    """
    (denominazione, P.IVA, codice fiscale) della controparte: Cliente per le
    fatture emesse, Fornitore per le ricevute. L'AdE usa il suffisso del ruolo
    corrispondente nei nomi campo, si prova entrambi e si usa quello valorizzato.

    La LISTA è la fonte primaria: è sempre disponibile anche quando il dettaglio
    fallisce/è temporaneamente vuoto, e per le persone fisiche riporta nome e
    cognome già uniti (nel dettaglio sono invece due campi separati, spesso con
    `denominazione*` vuoto, vedi `_nome_persona`). P.IVA o CF possono essere
    presenti singolarmente: nessuno dei due è garantito.
    """
    for suffisso in ("Cliente", "Fornitore"):
        piva = str(_campo(voce, f"piva{suffisso}", default="") or "").strip()
        cf = str(_campo(voce, f"cf{suffisso}", default="") or "").strip()
        denom = str(_campo(voce, f"denominazione{suffisso}", default="") or "").strip()
        if not (piva or cf or denom):
            continue
        if not denom and dettaglio:
            denom = (str(_campo(dettaglio, f"denominazione{suffisso}", default="") or "")
                     .strip() or _nome_persona(dettaglio, suffisso))
        return denom, piva, cf
    # Voce senza alcun dato di controparte (raro): ripiego sul solo dettaglio.
    if dettaglio:
        for suffisso in ("Cliente", "Fornitore"):
            piva = str(_campo(dettaglio, f"piva{suffisso}", default="") or "").strip()
            cf = str(_campo(dettaglio, f"cf{suffisso}", default="") or "").strip()
            denom = (str(_campo(dettaglio, f"denominazione{suffisso}", default="") or "")
                     .strip() or _nome_persona(dettaglio, suffisso))
            if piva or cf or denom:
                return denom, piva, cf
    return "", "", ""


def _estrai_riga(voce: dict, dettaglio: dict | None) -> dict:
    """
    Riga Excel per una fattura: unisce i campi della voce di lista con le righe di
    riepilogo IVA del dettaglio. Ritorna un dict con le chiavi di COLONNE_FISSE,
    più `iva` = {chiave_aliquota: [imponibile, imposta]} e i tre totali numerici.
    Per le note di credito tutti gli importi sono negati.
    """
    intestazione = dettaglio if isinstance(dettaglio, dict) else {}
    tipo_doc = str(_campo(voce, "tipoDocumento", "tipoDoc", "descTipoDocumento",
                          default=_campo(intestazione, "tipoDocumento", "tipoDoc",
                                         default="")) or "")
    denom, piva, cf = _dati_controparte(voce, dettaglio)
    id_invio = (_campo(intestazione, "idInvio", default="")
               or _campo(voce.get("fileDownload") or {}, "idInvio", default="")
               or _campo(voce, "idInvio", "idFattura", default=""))
    riga = {
        "Data": _data_it(_campo(voce, "dataFattura", "dataEmissione", "data",
                                "dataRicezione", default=_campo(
                                    intestazione, "dataFattura", default=""))),
        "N. Fattura": _campo(voce, "numeroFattura", "numero", "numFattura",
                             default="") or "",
        "ID Invio": str(id_invio or ""),
        "Tipo Documento": tipo_doc,
        "Cliente/Fornitore": denom,
        "Partita IVA": piva,
        "Codice Fiscale": cf,
        "Bollo Virtuale": _campo(voce, "bolloVirtuale", "bollo",
                                 default=_campo(intestazione, "bolloVirtuale",
                                               default="")) or "",
    }

    segno = -1.0 if _e_nota_credito(tipo_doc) else 1.0
    iva: dict[str, list[float]] = {}
    tot_imp = tot_iva = 0.0
    for riga_iva in _righe_riepilogo_iva(dettaglio):
        chiave = _chiave_aliquota(riga_iva)
        if chiave is None:
            continue
        imponibile = segno * _importo(_campo(riga_iva, "imponibile",
                                             "imponibileImporto"))
        imposta = segno * _importo(_campo(riga_iva, "imposta", "impostaImporto"))
        coppia = iva.setdefault(chiave, [0.0, 0.0])
        coppia[0] += imponibile
        coppia[1] += imposta
        tot_imp += imponibile
        tot_iva += imposta

    riga["iva"] = iva
    riga["Tot. Imponibile"] = tot_imp
    riga["Tot. IVA"] = tot_iva
    riga["Totale Fattura"] = tot_imp + tot_iva
    return riga


def _ordina_chiavi_pivot(chiavi) -> list[str]:
    """Aliquote numeriche crescenti prima ("4%", "10%", "22%"), poi nature alfabetiche."""
    def _ordine(chiave: str):
        if chiave.endswith("%"):
            return (0, _importo(chiave[:-1]), "")
        return (1, 0.0, chiave)
    return sorted(chiavi, key=_ordine)


# ─────────────────────────────────────────────────────────────────────────────
# Scrittura xlsx (openpyxl, import lazy: dipendenza opzionale ruolo "excel")
# ─────────────────────────────────────────────────────────────────────────────

def _scrivi_xlsx(percorso: str, colonne_fisse: list[str], chiavi_pivot: list[str],
                 colonne_num: list[str], colonne_testo: list[str],
                 righe: list[dict], titolo: str = "Elenco") -> None:
    """
    Scrive un elenco pivot in `percorso`: colonne fisse (testo), coppie
    `Imp. X | IVA X` per ogni chiave pivot, colonne numeriche di coda (sommate
    nella riga TOTALI verde) e colonne di testo finali (escluse dai totali).
    Ogni riga è un dict con le chiavi delle colonne + `iva` = {chiave: (imp, iva)}.
    Intestazioni in grassetto, formato numerico contabile, prima riga bloccata.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        import fec_deps
        raise DownloadError("L'export Excel richiede il pacchetto «openpyxl»: "
                            f"{fec_deps.pip_install_hint(['openpyxl'])}")

    intestazioni = list(colonne_fisse)
    for chiave in chiavi_pivot:
        intestazioni += [f"Imp. {chiave}", f"IVA {chiave}"]
    intestazioni += list(colonne_num) + list(colonne_testo)

    wb = Workbook()
    ws = wb.active
    ws.title = titolo
    ws.append(intestazioni)
    for cella in ws[1]:
        cella.font = Font(bold=True)
    ws.freeze_panes = "A2"

    prima_num = len(colonne_fisse) + 1                       # prima colonna numerica (1-based)
    ultima_num = len(intestazioni) - len(colonne_testo)      # colonne di testo finali escluse

    totali = [0.0] * (ultima_num - prima_num + 1)
    for riga in righe:
        valori = [riga.get(c, "") for c in colonne_fisse]
        for chiave in chiavi_pivot:
            imp, iva = riga["iva"].get(chiave, (None, None))
            valori += [imp, iva]
        valori += [riga.get(c) for c in colonne_num]
        valori += [riga.get(c, "") for c in colonne_testo]
        ws.append(valori)
        for i, v in enumerate(valori[prima_num - 1:ultima_num]):
            if isinstance(v, (int, float)):
                totali[i] += v

    riga_tot = ["TOTALI"] + [""] * (prima_num - 2) + totali + [""] * len(colonne_testo)
    ws.append(riga_tot)
    verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    for cella in ws[ws.max_row]:
        cella.font = Font(bold=True)
        cella.fill = verde

    for col in range(prima_num, ultima_num + 1):
        for cella in ws[get_column_letter(col)][1:]:
            cella.number_format = "#,##0.00"
    for col, titolo in enumerate(intestazioni, start=1):
        larghezza = 22 if titolo == "Cliente/Fornitore" else max(12, len(titolo) + 3)
        ws.column_dimensions[get_column_letter(col)].width = larghezza

    wb.save(percorso)


# ─────────────────────────────────────────────────────────────────────────────
# Funzione pubblica
# ─────────────────────────────────────────────────────────────────────────────

def elenco_fatture_excel(auth: AuthResult, cf_cliente: str, piva: str, tipo: str,
                         dal: str, al: str, *, dest_dir: str | None = None,
                         sottocartella: bool = True,
                         control: "Controllo | None" = None, log=print) -> str:
    """
    Genera l'Excel «Elenco fatture» per il periodo [dal, al] (GGMMAAAA) e ritorna
    il percorso del file .xlsx creato. Periodi > 3 mesi spezzati internamente
    (guardia 12 mesi di spezza_periodo); un solo file in output.
    """
    _, etichetta = TIPI_ELENCO.get(tipo, (None, tipo))
    blocchi = spezza_periodo(dal, al, "%d%m%Y")
    log(f"Elenco fatture {etichetta} per {cf_cliente}  ({dal} -> {al})"
        + (f"  [{len(blocchi)} blocchi]" if len(blocchi) > 1 else ""))

    voci: list[dict] = []
    for b_dal, b_al in blocchi:
        if control:
            control.check()
        if len(blocchi) > 1:
            log(f"  Blocco {b_dal} -> {b_al}: richiedo l'elenco...")
        voci.extend(_lista_fatture(auth, tipo, b_dal, b_al))
    log(f"Trovate {len(voci)} fatture nell'intervallo.")
    if not voci:
        raise NessunDato("Nessuna fattura trovata nell'intervallo richiesto: "
                         "nessun file Excel generato.")

    righe: list[dict] = []
    for i, voce in enumerate(voci, start=1):
        if control:
            control.check()
        fattura_file = f"{voce.get('tipoInvio', '')}{voce.get('idFattura', '')}"
        dettaglio = _dettaglio_fattura(auth, fattura_file, log=log)
        righe.append(_estrai_riga(voce, dettaglio))
        if i % 10 == 0 or i == len(voci):
            log(f"  Dettaglio fatture: {i}/{len(voci)}")

    chiavi_pivot = _ordina_chiavi_pivot({k for r in righe for k in r["iva"]})

    cartella = _cartella(dest_dir, "ElencoFatture", cf_cliente, sottocartella)
    prefisso = (piva or cf_cliente).strip()
    percorso = os.path.join(cartella, f"{prefisso}_{dal}-{al}_{etichetta}.xlsx")
    _scrivi_xlsx(percorso, COLONNE_FISSE, chiavi_pivot,
                 ["Tot. Imponibile", "Tot. IVA", "Totale Fattura"],
                 ["Bollo Virtuale"], righe, titolo="Elenco fatture")
    log(f"\nElenco fatture generato ({len(righe)} righe, "
        f"{len(chiavi_pivot)} aliquote/nature).")
    log(f"File: {percorso}")
    return percorso


# ─────────────────────────────────────────────────────────────────────────────
# Corrispettivi - endpoint:
#   sintesi:   rs/corrispettivi/sintesi/dal/{dal}/al/{al}/piva/{piva}
#   elenco:    rs/corrispettivi/sintesi/elenco/dal/{dal}/al/{al}/piva/{piva}
#                  /tipoCorrispettivo/{RT|MC|DA}    -> {"corrispettivi": [...]}
#   dettaglio: rs/corrispettivi/dettaglio/{tipo}{idInvio}
#                  -> {"datiContabiliRT_MC": [righe IVA], ...totali}
# Date GGMMAAAA; la P.IVA è OBBLIGATORIA (fa parte dell'URL).
# ─────────────────────────────────────────────────────────────────────────────

# tipoCorrispettivo → campo conteggio della sintesi. Solo RT è confermato dalla
# cattura; MC/DA sono dedotti dai partial HTML della SPA (elencoinvii-rt-mc /
# colonnine / carb) e comunque interrogati solo se la sintesi ne conta invii.
TIPI_CORRISPETTIVO = {
    "RT": "registratoriInvii",      # registratori telematici (confermato)
    "MC": "multicassaInvii",        # multicassa (dedotto)
    "DA": "distributoriInvii",      # distributori automatici (dedotto)
}

# Colonne fisse dell'Excel corrispettivi (formato dallo userscript FE-Utility).
COLONNE_FISSE_CORR = ["ID Invio", "Data", "Tipo dispositivo"]


def _sintesi_corrispettivi(auth: AuthResult, dal: str, al: str, piva: str) -> dict:
    """Sintesi del periodo: conteggio invii per tipo dispositivo (dict JSON)."""
    url = (f"{IVASERVIZI}/cons/cons-services/rs/corrispettivi/sintesi"
           f"/dal/{dal}/al/{al}/piva/{piva}?v={unix_time()}")
    try:
        r = auth.session.get(url, headers=auth.headers, verify=False, timeout=HTTP_TIMEOUT)
    except requests.exceptions.RequestException as exc:
        raise DownloadError(f"Nessuna risposta dall'AdE per la sintesi corrispettivi: {exc}.")
    if r.status_code != 200:
        raise DownloadError(
            f"La sintesi corrispettivi è fallita (HTTP {r.status_code}). "
            "La sessione potrebbe essere scaduta o la P.IVA non appartenere all'utenza.")
    try:
        return r.json()
    except ValueError:
        raise DownloadError("Risposta sintesi corrispettivi non in formato JSON.")


def _lista_corrispettivi(auth: AuthResult, dal: str, al: str, piva: str,
                         tipo_corr: str) -> list[dict]:
    """Elenco invii del periodo per un tipoCorrispettivo (voci JSON grezze)."""
    url = (f"{IVASERVIZI}/cons/cons-services/rs/corrispettivi/sintesi/elenco"
           f"/dal/{dal}/al/{al}/piva/{piva}/tipoCorrispettivo/{tipo_corr}"
           f"?v={unix_time()}")
    try:
        r = auth.session.get(url, headers=auth.headers, verify=False, timeout=HTTP_TIMEOUT)
    except requests.exceptions.RequestException as exc:
        raise DownloadError(f"Nessuna risposta dall'AdE per l'elenco corrispettivi: {exc}.")
    if r.status_code != 200:
        raise DownloadError(f"L'elenco corrispettivi {tipo_corr} è fallito "
                            f"(HTTP {r.status_code}).")
    try:
        data = r.json()
    except ValueError:
        raise DownloadError("Risposta elenco corrispettivi non in formato JSON.")
    return data.get("corrispettivi") or []


def _dettaglio_corrispettivo(auth: AuthResult, tipo_corr: str,
                             id_invio: str) -> dict | None:
    """Dettaglio JSON di un invio (righe IVA). None su errore (riga coi soli totali)."""
    url = (f"{IVASERVIZI}/cons/cons-services/rs/corrispettivi/dettaglio"
           f"/{tipo_corr}{id_invio}?v={unix_time()}")
    try:
        r = auth.session.get(url, headers=auth.headers, verify=False, timeout=HTTP_TIMEOUT)
    except requests.exceptions.RequestException:
        return None
    if r.status_code != 200:
        return None
    try:
        return r.json()
    except ValueError:
        return None


def _chiave_aliquota_corr(riga_iva: dict) -> str:
    """
    Chiave pivot di una riga contabile corrispettivi (dallo userscript FE-Utility):
    ventilazione → «Ventilazione IVA»; aliquota numerica → «X%»; natura → codice;
    altrimenti «Esente/N.I.».
    """
    if str(_campo(riga_iva, "ventilazione", default="") or "").strip():
        return "Ventilazione IVA"
    aliquota = _campo(riga_iva, "aliquota", "aliquotaIVA", "aliquotaIva")
    if aliquota is not None and _importo(aliquota) > 0:
        return f"{_importo(aliquota):g}%".replace(".", ",")
    natura = str(_campo(riga_iva, "natura", "codiceNatura", default="") or "").strip()
    if natura:
        return natura.upper()
    return "Esente/N.I."


def _estrai_riga_corr(voce: dict, dettaglio: dict | None) -> dict:
    """
    Riga Excel per un invio corrispettivi: dati dell'elenco + righe contabili del
    dettaglio (`datiContabiliRT_MC`). Senza dettaglio usa i totali dell'elenco.
    """
    riga = {
        "ID Invio": str(_campo(voce, "idInvio", default="") or ""),
        "Data": _data_it(_campo(voce, "timeRilevazione", "dataAccoglienzaFile")),
        "Tipo dispositivo": str(_campo(voce, "tipoDispositivo", default="") or ""),
        "_matricola": str(_campo(voce, "matricolaDispositivo", default="") or ""),
        "_ordinamento": str(_campo(voce, "timeRilevazione", default="") or ""),
    }

    righe_iva = _righe_riepilogo_iva(dettaglio) if dettaglio else []
    iva: dict[str, list[float]] = {}
    tot_imp = tot_iva = 0.0
    for riga_iva in righe_iva:
        chiave = _chiave_aliquota_corr(riga_iva)
        imponibile = _importo(_campo(riga_iva, "imponibile", "imponibileImporto"))
        imposta = _importo(_campo(riga_iva, "imposta", "impostaImporto"))
        coppia = iva.setdefault(chiave, [0.0, 0.0])
        coppia[0] += imponibile
        coppia[1] += imposta
        tot_imp += imponibile
        tot_iva += imposta
    if not righe_iva:                       # fallback sui totali dell'elenco
        tot_imp = _importo(_campo(voce, "ammontareTotale", "importoParzialeTotale"))
        tot_iva = _importo(_campo(voce, "impostaTotale"))

    fonte_memo = dettaglio if righe_iva else voce
    riga["iva"] = iva
    riga["Tot. Imponibile"] = tot_imp
    riga["Tot. IVA"] = tot_iva
    riga["Resi"] = _importo(_campo(fonte_memo, "resi"))
    riga["Annulli"] = _importo(_campo(fonte_memo, "annullato", "annulli"))
    riga["Totale Corrispettivi"] = tot_imp + tot_iva
    return riga


def elenco_corrispettivi_excel(auth: AuthResult, cf_cliente: str, piva: str,
                               dal: str, al: str, *, dest_dir: str | None = None,
                               sottocartella: bool = True,
                               control: "Controllo | None" = None,
                               log=print) -> list[str]:
    """
    Genera gli Excel «Elenco corrispettivi» per il periodo [dal, al] (GGMMAAAA):
    **un file per matricola dispositivo** (come lo userscript FE-Utility), una riga
    per invio, pivot per aliquota/ventilazione/natura, Resi/Annulli a memo (esclusi
    dal Totale Corrispettivi). La P.IVA è obbligatoria (fa parte degli URL AdE).
    Ritorna l'elenco dei percorsi creati.
    """
    piva = (piva or "").strip()
    if not piva:
        raise DownloadError("Per i corrispettivi la P.IVA è obbligatoria (fa parte "
                            "dell'indirizzo delle chiamate AdE) e non è stato possibile "
                            "ricavarla dall'utenza di lavoro attiva: indicala a mano.")
    blocchi = spezza_periodo(dal, al, "%d%m%Y")
    log(f"Elenco corrispettivi per P.IVA {piva}  ({dal} -> {al})"
        + (f"  [{len(blocchi)} blocchi]" if len(blocchi) > 1 else ""))

    # Sintesi per capire quali tipi dispositivo hanno invii nel periodo.
    voci: list[tuple[str, dict]] = []       # (tipo_corr, voce elenco)
    for b_dal, b_al in blocchi:
        if control:
            control.check()
        sintesi = _sintesi_corrispettivi(auth, b_dal, b_al, piva)
        for tipo_corr, campo_conteggio in TIPI_CORRISPETTIVO.items():
            n = int(_importo(_campo(sintesi, campo_conteggio, default=0)))
            if n <= 0:
                continue
            log(f"  {b_dal} -> {b_al}: {n} invii {tipo_corr} "
                f"({campo_conteggio}); scarico l'elenco...")
            for voce in _lista_corrispettivi(auth, b_dal, b_al, piva, tipo_corr):
                voci.append((tipo_corr, voce))
        # Tipi presenti in sintesi ma non mappati (es. carburanti): segnalo e proseguo.
        for campo in ("carburantiInvii", "documentiCommerciali", "registratoriCassa",
                      "torretteEnergia"):
            n = int(_importo(_campo(sintesi, campo, default=0)))
            if n > 0:
                log(f"   ⚠️  {n} invii «{campo}» non supportati dall'export "
                    "(tipoCorrispettivo non ancora noto): esclusi.")

    log(f"Trovati {len(voci)} invii nell'intervallo.")
    if not voci:
        raise NessunDato("Nessun corrispettivo trovato nell'intervallo richiesto: "
                         "nessun file Excel generato.")

    righe: list[dict] = []
    for i, (tipo_corr, voce) in enumerate(voci, start=1):
        if control:
            control.check()
        id_invio = str(_campo(voce, "idInvio", default="") or "")
        dettaglio = None
        if _campo(voce, "dettaglio", default=False):
            dettaglio = _dettaglio_corrispettivo(auth, tipo_corr, id_invio)
            if dettaglio is None:
                log(f"   ⚠️  Dettaglio {tipo_corr}{id_invio} non ottenuto: riga "
                    "coi soli totali di elenco.")
        righe.append(_estrai_riga_corr(voce, dettaglio))
        if i % 10 == 0 or i == len(voci):
            log(f"  Dettaglio invii: {i}/{len(voci)}")

    # Un file per matricola dispositivo.
    per_matricola: dict[str, list[dict]] = {}
    for riga in righe:
        per_matricola.setdefault(riga["_matricola"] or "senza_matricola",
                                 []).append(riga)

    cartella = _cartella(dest_dir, "ElencoCorrispettivi", cf_cliente, sottocartella)
    percorsi: list[str] = []
    for matricola, righe_m in sorted(per_matricola.items()):
        righe_m.sort(key=lambda r: r["_ordinamento"])
        chiavi_pivot = _ordina_chiavi_pivot({k for r in righe_m for k in r["iva"]})
        percorso = os.path.join(cartella,
                                f"{piva}_{dal}-{al}_corrispettivi_{matricola}.xlsx")
        _scrivi_xlsx(percorso, COLONNE_FISSE_CORR, chiavi_pivot,
                     ["Tot. Imponibile", "Tot. IVA", "Resi", "Annulli",
                      "Totale Corrispettivi"],
                     [], righe_m, titolo="Elenco corrispettivi")
        log(f"File ({len(righe_m)} invii, matricola {matricola}): {percorso}")
        percorsi.append(percorso)

    log(f"\nGenerati {len(percorsi)} file (uno per matricola dispositivo).")
    return percorsi


# ─────────────────────────────────────────────────────────────────────────────
# Discovery (dev-only): dump dei JSON reali per confermare i nomi campo
# ─────────────────────────────────────────────────────────────────────────────

def dump_json_esempio(auth: AuthResult, tipo: str, dal: str, al: str,
                      dest_dir: str = "_materiale", log=print) -> None:
    """
    Salva in `dest_dir` la lista fatture del PRIMO blocco del periodo e il dettaglio
    della PRIMA fattura (`dump_lista_{tipo}.json` / `dump_dettaglio_{tipo}.json`).
    Serve solo in sviluppo, per confermare i nomi dei campi JSON (la lista espone di
    certo solo tipoInvio/idFattura; il resto va verificato sul vivo). ⚠️ I dump possono
    contenere dati personali reali: tenerli in `_materiale/` (fuori da Git).
    """
    b_dal, b_al = spezza_periodo(dal, al, "%d%m%Y")[0]
    voci = _lista_fatture(auth, tipo, b_dal, b_al)
    os.makedirs(dest_dir, exist_ok=True)
    percorso = os.path.join(dest_dir, f"dump_lista_{tipo}.json")
    with open(percorso, "w", encoding="utf-8") as fh:
        json.dump(voci, fh, indent=2, ensure_ascii=False)
    log(f"Lista ({len(voci)} voci) salvata in {percorso}")
    if not voci:
        log("Nessuna fattura nel blocco: dettaglio non scaricabile.")
        return
    fattura_file = f"{voci[0].get('tipoInvio', '')}{voci[0].get('idFattura', '')}"
    dettaglio = _dettaglio_fattura(auth, fattura_file, log=log)
    percorso = os.path.join(dest_dir, f"dump_dettaglio_{tipo}.json")
    with open(percorso, "w", encoding="utf-8") as fh:
        json.dump(dettaglio, fh, indent=2, ensure_ascii=False)
    log(f"Dettaglio fattura {fattura_file} salvato in {percorso}")


def cattura_har_debug(nomeutente: str, pin: str, password: str,
                      capture_dir: str = "_materiale", log=print) -> tuple[str, str]:
    """
    Cattura HAR a navigazione libera per il debug di fatture e corrispettivi
    (utile ad esempio per scoprire nuovi endpoint JSON).

    Wrapper sopra lo strumento generico `ade_auth.cattura_har_navigazione`
    (riusato anche dall'investigazione sulle deleghe, dalla tab Test Login),
    qui invocato col prefisso file `capture_debug_`.
    """
    from ade_auth import cattura_har_navigazione, AuthError
    try:
        return cattura_har_navigazione(nomeutente, pin, password,
                                       capture_dir=capture_dir,
                                       prefisso="capture_debug", log=log)
    except AuthError as exc:
        raise DownloadError(str(exc))


if __name__ == "__main__":
    # Uso dev:  python fec_utility.py --tipo emesse --dal 01012026 --al 31032026 \
    #               --cf-cliente 12345678901 [--profilo 1] [--backend requests]
    # Credenziali: cf/pin/cfstudio da fec_store (come la GUI); password da
    # variabile d'ambiente FEC_PASSWORD o richiesta a video (mai salvata su disco).
    import argparse
    import getpass

    import ade_auth
    import fec_store

    ap = argparse.ArgumentParser(description="Dump JSON di discovery per l'export Excel (dev)")
    ap.add_argument("--tipo", default="emesse", choices=sorted(TIPI_ELENCO))
    ap.add_argument("--dal", required=True, help="data inizio GGMMAAAA")
    ap.add_argument("--al", required=True, help="data fine GGMMAAAA")
    ap.add_argument("--cf-cliente", required=True)
    ap.add_argument("--profilo", type=int, default=1)
    ap.add_argument("--backend", default="requests", choices=["requests", "browser"])
    args = ap.parse_args()

    cred = fec_store.load_credentials()
    if not cred:
        raise SystemExit("Nessuna credenziale salvata (fec_credentials.dat): "
                         "salva le credenziali dalla GUI prima di usare il dump.")
    password = os.environ.get("FEC_PASSWORD") or getpass.getpass("Password Entratel: ")
    creds = ade_auth.Creds(nomeutente=cred.get("cf", ""), pin=cred.get("pin", ""),
                           password=password, cfstudio=cred.get("cfstudio", ""),
                           cf_cliente=args.cf_cliente, profilo=args.profilo)
    auth = ade_auth.autentica(creds, backend=args.backend)
    dump_json_esempio(auth, args.tipo, args.dal, args.al)
